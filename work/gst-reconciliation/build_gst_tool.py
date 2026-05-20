"""
Builds gst-reconciliation-tool.xlsx — a working GST input-credit reconciliation tool.

It matches a Purchase Register (your books) against GSTR-2B (what suppliers filed),
on a key of  GSTIN | Invoice No, and flags every line as one of:
  - Matched
  - Matched with difference   (value/tax mismatch)
  - In books, not in 2B        (ITC AT RISK — supplier hasn't filed)
  - In 2B, not in books        (missing entry in your books, or not yours)

All numbers below are DUMMY data, designed to show each case. The reconciliation
is done with LIVE Excel formulas (INDEX/MATCH, COUNTIF, SUMIF) so it recalculates
if you change the inputs. Run:  python3 build_gst_tool.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ---- styling helpers --------------------------------------------------------
INK   = "15130E"
BLUE  = "244B73"
PAPER = "F6F2E9"
hdr_fill   = PatternFill("solid", fgColor=BLUE)
title_font = Font(bold=True, size=14, color=INK)
hdr_font   = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="C8C0AE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
money = '#,##0'

green = PatternFill("solid", fgColor="DDE8E0")
amber = PatternFill("solid", fgColor="F4E9D2")
red   = PatternFill("solid", fgColor="F3E3DD")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

# ---- dummy data -------------------------------------------------------------
# Purchase Register (your books): Supplier, GSTIN, InvoiceNo, Date, Taxable, IGST, CGST, SGST
books = [
    ("Alpha Traders",  "27AABCA1234A1Z5", "INV-1001", "2026-04-03", 100000, 18000,    0,    0),
    ("Bharat Supplies","33BBPCB5678B1Z3", "INV-2002", "2026-04-07",  50000,     0, 4500, 4500),
    ("Chola Steel",    "33CCDCC9012C1Z1", "INV-3003", "2026-04-12",  80000, 14400,    0,    0),
    ("Deccan Pipes",   "36DDECD3456D1Z9", "INV-4004", "2026-04-15",  30000,  5400,    0,    0),
    ("Everest Tools",  "29EEFCE7890E1Z7", "INV-5005", "2026-04-18",  25000,     0, 2250, 2250),
    ("Falcon Pack",    "24FFGCF2345F1Z2", "INV-6006", "2026-04-22",  60000, 10800,    0,    0),
    ("Ganga Chem",     "33GGHCG6789G1Z4", "INV-7007", "2026-04-25",  45000,     0, 4050, 4050),
    ("Hind Logistics", "27HHICH0123H1Z6", "INV-8008", "2026-04-28",  20000,  3600,    0,    0),
]
# GSTR-2B (what suppliers actually filed). Note vs books:
#  - 3003 has a DIFFERENT value (78000/14040)  -> "Matched with difference"
#  - 4004 and 7007 are MISSING                 -> "In books, not in 2B (ITC at risk)"
#  - 9009 is EXTRA, not in books               -> "In 2B, not in books"
gstr2b = [
    ("Alpha Traders",  "27AABCA1234A1Z5", "INV-1001", "2026-04-03", 100000, 18000,    0,    0),
    ("Bharat Supplies","33BBPCB5678B1Z3", "INV-2002", "2026-04-07",  50000,     0, 4500, 4500),
    ("Chola Steel",    "33CCDCC9012C1Z1", "INV-3003", "2026-04-12",  78000, 14040,    0,    0),
    ("Everest Tools",  "29EEFCE7890E1Z7", "INV-5005", "2026-04-18",  25000,     0, 2250, 2250),
    ("Falcon Pack",    "24FFGCF2345F1Z2", "INV-6006", "2026-04-22",  60000, 10800,    0,    0),
    ("Hind Logistics", "27HHICH0123H1Z6", "INV-8008", "2026-04-28",  20000,  3600,    0,    0),
    ("Iyer Agencies",  "33IIJCI4567I1Z8", "INV-9009", "2026-04-29",  40000,  7200,    0,    0),
]

DATA_HEADERS = ["Supplier", "Supplier GSTIN", "Invoice No", "Invoice Date",
                "Taxable Value", "IGST", "CGST", "SGST", "Total Tax", "Match Key"]

wb = openpyxl.Workbook()

# ============================ Purchase Register ============================
def write_data_sheet(name, title, rows, extra_cols):
    ws = wb.create_sheet(name)
    ws["A1"] = title; ws["A1"].font = title_font
    headers = DATA_HEADERS + extra_cols
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    r = 3
    for (sup, gstin, inv, date, tax, igst, cgst, sgst) in rows:
        ws.cell(r, 1, sup); ws.cell(r, 2, gstin); ws.cell(r, 3, inv); ws.cell(r, 4, date)
        ws.cell(r, 5, tax).number_format = money
        ws.cell(r, 6, igst).number_format = money
        ws.cell(r, 7, cgst).number_format = money
        ws.cell(r, 8, sgst).number_format = money
        ws.cell(r, 9, f"=F{r}+G{r}+H{r}").number_format = money   # Total Tax
        ws.cell(r, 10, f'=B{r}&"|"&C{r}')                          # Match Key
        r += 1
    ws.freeze_panes = "A3"
    widths = [16, 18, 12, 13, 14, 10, 9, 9, 11, 22] + [26] * len(extra_cols)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws, r - 1  # last data row

books_ws, books_last = write_data_sheet(
    "Purchase_Register", "PURCHASE REGISTER  (your books)", books,
    ["2B Taxable", "2B Total Tax", "Tax Difference", "Status"])

twob_ws, twob_last = write_data_sheet(
    "GSTR_2B", "GSTR-2B  (what suppliers filed)", gstr2b, ["In Books?", "Status"])

# ---- Purchase Register: reconciliation formulas (cols K..N) ----
for r in range(3, books_last + 1):
    # 2B Taxable  (K) and 2B Total Tax (L) pulled by match key
    books_ws.cell(r, 11, f'=IFERROR(INDEX(GSTR_2B!$E:$E,MATCH(J{r},GSTR_2B!$J:$J,0)),"--")').number_format = money
    books_ws.cell(r, 12, f'=IFERROR(INDEX(GSTR_2B!$I:$I,MATCH(J{r},GSTR_2B!$J:$J,0)),"--")').number_format = money
    books_ws.cell(r, 13, f'=IF(ISNUMBER(L{r}),I{r}-L{r},"--")').number_format = money
    books_ws.cell(r, 14,
        f'=IF(NOT(ISNUMBER(L{r})),"In books, not in 2B (ITC at risk)",'
        f'IF(AND(ABS(I{r}-L{r})<=1,ABS(E{r}-K{r})<=1),"Matched","Matched with difference"))')

# ---- GSTR-2B: is each 2B invoice present in books? ----
for r in range(3, twob_last + 1):
    twob_ws.cell(r, 11, f'=IF(COUNTIF(Purchase_Register!$J:$J,J{r})>0,"Yes","No")')
    twob_ws.cell(r, 12, f'=IF(K{r}="No","In 2B, not in books","Also in books")')

# ---- conditional formatting (colour the Status columns) ----
def colour_status(ws, col_letter, last):
    rng = f"{col_letter}3:{col_letter}{last}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("risk",{col_letter}3))'], fill=red))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("not in books",{col_letter}3))'], fill=red))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("difference",{col_letter}3))'], fill=amber))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'EXACT({col_letter}3,"Matched")'], fill=green))

colour_status(books_ws, "N", books_last)
colour_status(twob_ws, "L", twob_last)

# ============================ Summary ============================
sm = wb.create_sheet("Summary")
sm["A1"] = "RECONCILIATION SUMMARY"; sm["A1"].font = title_font
rows = [
    ("Category", "Count", "Tax amount (Rs)"),
    ('Matched', '=COUNTIF(Purchase_Register!N:N,"Matched")',
        '=SUMIF(Purchase_Register!N:N,"Matched",Purchase_Register!I:I)'),
    ('Matched with difference', '=COUNTIF(Purchase_Register!N:N,"Matched with difference")',
        '=SUMIF(Purchase_Register!N:N,"Matched with difference",Purchase_Register!I:I)'),
    ('In books, not in 2B (ITC AT RISK)', '=COUNTIF(Purchase_Register!N:N,"In books, not in 2B (ITC at risk)")',
        '=SUMIF(Purchase_Register!N:N,"In books, not in 2B (ITC at risk)",Purchase_Register!I:I)'),
    ('In 2B, not in books', '=COUNTIF(GSTR_2B!L:L,"In 2B, not in books")',
        '=SUMIF(GSTR_2B!L:L,"In 2B, not in books",GSTR_2B!I:I)'),
    ('', '', ''),
    ('Total ITC in books', '=COUNT(Purchase_Register!I3:I100)', '=SUM(Purchase_Register!I:I)'),
    ('ITC safe to claim now (matched)', '',
        '=SUMIF(Purchase_Register!N:N,"Matched",Purchase_Register!I:I)'),
    ('ITC to follow up (at risk + difference)', '',
        '=SUM(Purchase_Register!I:I)-SUMIF(Purchase_Register!N:N,"Matched",Purchase_Register!I:I)'),
]
for i, (a, b, c) in enumerate(rows, 2):
    sm.cell(i, 1, a); sm.cell(i, 2, b); sm.cell(i, 3, c)
    if c.startswith("=SUM") or c.startswith("=SUMIF"):
        sm.cell(i, 3).number_format = money
style_header(sm, 2, 3)
sm.column_dimensions["A"].width = 38
sm.column_dimensions["B"].width = 10
sm.column_dimensions["C"].width = 18

# ============================ ReadMe ============================
rm = wb.create_sheet("ReadMe")
rm.column_dimensions["A"].width = 100
notes = [
    ("GST INPUT-CREDIT RECONCILIATION TOOL", title_font),
    ("", None),
    ("WHAT IT DOES: matches your Purchase Register against GSTR-2B on (GSTIN | Invoice No) and", None),
    ("flags each line. Replace the dummy data in the two input sheets with your own and it recomputes.", None),
    ("", None),
    ("THE FOUR OUTCOMES:", Font(bold=True)),
    ("  - Matched .................. present in both, values agree -> ITC safe to claim.", None),
    ("  - Matched with difference .. present in both, value/tax differs -> investigate & fix.", None),
    ("  - In books, not in 2B ...... you recorded it, supplier hasn't filed -> ITC AT RISK, follow up.", None),
    ("  - In 2B, not in books ...... appears in 2B but not in your books -> missed entry, or not yours.", None),
    ("", None),
    ("WHY IT MATTERS (ITC rules, in brief):", Font(bold=True)),
    ("  - Sec 16(2)(aa) / Rule 36(4): ITC can be claimed only if the invoice appears in GSTR-2B.", None),
    ("    So 'In books, not in 2B' credit cannot be taken until the supplier files it.", None),
    ("  - Other conditions for ITC: valid tax invoice held, goods/services received, tax actually", None),
    ("    paid to government, and the claim is made within the time limit (by 30 Nov of next FY).", None),
    ("  - Sec 17(5): some credits are blocked (e.g. motor vehicles, personal use) - check manually.", None),
    ("", None),
    ("NOTE: dummy data for demonstration. Verify the current rules before relying on this in practice.", Font(italic=True)),
]
for i, (text, font) in enumerate(notes, 1):
    cell = rm.cell(i, 1, text)
    if font: cell.font = font

# order sheets, drop default
wb.remove(wb["Sheet"])
wb.move_sheet("ReadMe", -(len(wb.sheetnames)))   # ReadMe first
wb._sheets.sort(key=lambda s: ["ReadMe", "Purchase_Register", "GSTR_2B", "Summary"].index(s.title))

wb.save("gst-reconciliation-tool.xlsx")
print("Wrote gst-reconciliation-tool.xlsx with sheets:", wb.sheetnames)
